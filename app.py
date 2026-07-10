import streamlit as st
import pandas as pd
import re
import io
import openpyxl
from copy import copy
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

# 페이지 설정
st.set_page_config(
    page_title="자동 발주 파일 생성기",
    page_icon="📦",
    layout="wide"
)

# ==========================================
# 설정 및 함수들
# ==========================================
MARKET_CONFIG = {
    'naver': {'key': '스마트스토어', 'skip': 1, 'order': 1},
    'coupang': {'key': 'DeliveryList', 'skip': 0, 'order': 2},
    'own': {'key': 'orders', 'skip': 0, 'order': 3},
    'esm': {'key': '신규주문', 'skip': 0, 'order': 4},
    '11st': {'key': 'allList', 'skip': 2, 'order': 5},
    '11st_manual': {'key': '11번가', 'skip': 0, 'order': 5},
    'wadiz': {'key': '발송 처리용 주문', 'skip': 0, 'order': 6}
}

NAVER_DELIVERY_TEMPLATE_NAME = "네이버_엑셀발송_양식.xls"
NAVER_DELIVERY_COLUMNS = ['상품주문번호', '배송방법', '택배사', '송장번호']
NAVER_DELIVERY_COLUMN_ALIASES = {
    '상품주문번호': ['상품주문번호'],
    '배송방법': ['배송방법', '배송 방법'],
    '택배사': ['택배사', '택배사명'],
    '송장번호': ['송장번호', '운송장번호']
}
NAVER_DELIVERY_METHOD = "택배,등기,소포"
NAVER_DELIVERY_COMPANY = "CJ대한통운"

def find_naver_delivery_template():
    for path in (
        Path("sample_data") / NAVER_DELIVERY_TEMPLATE_NAME,
        Path(NAVER_DELIVERY_TEMPLATE_NAME),
    ):
        if path.exists():
            return path
    return None

def clean_phone(phone):
    if pd.isna(phone): return ""
    return re.sub(r'[^0-9]', '', str(phone))

def normalize_excel_id(value):
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r'\d+\.0', text):
        return text[:-2]
    return text

def identify_product(name):
    name_str = str(name)
    name_upper = name_str.upper()
    name_lower = name_str.lower()
    
    # 리퍼 제품 여부 확인 (_Re 표기 또는 한글 '리퍼'/'리퍼제품')
    is_refurb = '_RE' in name_upper or '리퍼' in name_str

    # IH, OH, PH, SH 코드 우선 확인 (리퍼면 _Re 접미사 부착)
    for code in ('IH', 'OH', 'PH', 'SH'):
        if code in name_upper:
            return f"{code}_Re" if is_refurb else code
    
    # 기타 제품 매핑
    if '케이블s' in name_lower:
        return '케이블s'
    if '케이블' in name_str:
        if '스위치' in name_str:
            return '케이블s'
        else:
            return '케이블(일반)'
    if '거치대' in name_str or '휴대폰' in name_str:
        return '휴대폰거치대'
    if '번호판' in name_str or '차량번호' in name_str:
        return '차량번호판'
    if '망치' in name_str or '차량용망치' in name_str:
        return '차량용망치'
    if '도막' in name_str or '측정기' in name_str:
        return '도막측정기'

    return name

# 판매자/업체 상품코드(예: 'PH', 'SH_Re')를 표준 품목명으로 정규화
# 네이버 '판매자 상품코드', 쿠팡 '업체상품코드'처럼 코드가 명시된 컬럼을 우선 신뢰한다.
_CODE_RE = re.compile(r'^(IH|OH|PH|SH)(?:[ _\-]?(RE))?$')
def code_to_item(raw):
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = _CODE_RE.match(s.upper())
    if not m:
        return None
    return f"{m.group(1)}_Re" if m.group(2) else m.group(1)

def get_message(row, cols):
    for col in cols:
        if col in row and pd.notna(row[col]) and str(row[col]).strip() != "":
            return str(row[col]).strip()
    return ""

def pick_first_col(columns, candidates):
    for col in candidates:
        if col in columns:
            return col
    return None

def format_date(value):
    if pd.isna(value):
        return ""
    try:
        return pd.to_datetime(value).strftime('%Y.%m.%d')
    except Exception:
        return str(value)

def detect_market_by_columns(df):
    cols = set(df.columns.astype(str))

    # 와디즈 감지 (고유 컬럼)
    required_wadiz = {'주문 번호', '주문 상품', '주문 수량', '받는 분'}
    if required_wadiz.issubset(cols):
        return 'wadiz'

    required_11st = {'주문번호', '주소', '상품명', '수량'}
    name_cols_11st = {'수취인', '받는분'}
    phone_cols_11st = {'휴대폰번호', '수취인연락처'}
    if required_11st.issubset(cols) and cols.intersection(name_cols_11st) and cols.intersection(phone_cols_11st):
        return '11st_manual'

    return None

def sort_xlsx_preserving_format(file_content, target_col_name):
    """원본 서식을 유지하며 업체상품코드 기준으로 정렬"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        
        try:
            col_idx = header.index(target_col_name)
        except:
            return None

        rows = list(ws.iter_rows(min_row=2, values_only=False))
        rows.sort(key=lambda x: str(x[col_idx].value) if x[col_idx].value is not None else "")

        data_styles = []
        for row in rows:
            data_styles.append([(cell.value, cell._style) for cell in row])

        ws.delete_rows(2, ws.max_row)
        for r_idx, row_data in enumerate(data_styles, start=2):
            for c_idx, (val, style) in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if style:
                    cell._style = style
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        return None

def _set_text_format_for_columns(ws, header, target_cols=None, keyword_cols=None):
    target_cols = target_cols or []
    keyword_cols = keyword_cols or []
    col_indexes = set()

    for col_name in target_cols:
        if col_name in header:
            col_indexes.add(header.index(col_name) + 1)

    for idx, name in enumerate(header, start=1):
        name_str = str(name) if name is not None else ""
        if any(keyword in name_str for keyword in keyword_cols):
            col_indexes.add(idx)

    if not col_indexes:
        return

    for col_idx in col_indexes:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.value = str(cell.value)
            cell.number_format = '@'

def apply_text_format_to_excel_bytes(file_bytes, target_cols=None, keyword_cols=None):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        _set_text_format_for_columns(ws, header, target_cols=target_cols, keyword_cols=keyword_cols)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception:
        return file_bytes

def _read_tabular_file(file_content, file_name, skiprows=0):
    name = file_name.lower()
    if name.endswith('.csv'):
        return pd.read_csv(io.BytesIO(file_content), skiprows=skiprows)
    return pd.read_excel(io.BytesIO(file_content), skiprows=skiprows)

def _read_naver_order_df(file_content, file_name):
    for skiprows in (MARKET_CONFIG['naver']['skip'], 0):
        try:
            df = _read_tabular_file(file_content, file_name, skiprows=skiprows)
            df.columns = df.columns.astype(str).str.strip()
            if '상품주문번호' in df.columns:
                return df
        except Exception:
            continue
    return None

def _find_header_row(ws, required_header):
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        values = [cell.value for cell in ws[row_idx]]
        if required_header in values:
            return row_idx, values
    return 1, [cell.value for cell in ws[1]]

def _ensure_columns(ws, header_row_idx, header, columns):
    header = list(header)
    for col_name in columns:
        if col_name not in header:
            header.append(col_name)
            ws.cell(row=header_row_idx, column=len(header), value=col_name)
    return header

def _find_naver_delivery_header(header, canonical_name):
    for alias in NAVER_DELIVERY_COLUMN_ALIASES[canonical_name]:
        if alias in header:
            return alias
    return None

def _ensure_naver_delivery_columns(ws, header_row_idx, header):
    header = list(header)
    for col_name in NAVER_DELIVERY_COLUMNS:
        if _find_naver_delivery_header(header, col_name) is None:
            header.append(col_name)
            ws.cell(row=header_row_idx, column=len(header), value=col_name)
    return header

def _read_template_header(template_content, template_name):
    if not template_content or not template_name:
        return None
    try:
        preview = pd.read_excel(io.BytesIO(template_content), header=None, nrows=20)
        for _, row in preview.iterrows():
            values = [str(value).strip() if pd.notna(value) else None for value in row.tolist()]
            if '상품주문번호' in values:
                while values and values[-1] is None:
                    values.pop()
                return values
    except Exception:
        return None
    return None

def _write_naver_delivery_xlsx(rows, template_content=None, template_name=None):
    template_is_xlsx = template_content and template_name and template_name.lower().endswith('.xlsx')

    if template_is_xlsx:
        wb = openpyxl.load_workbook(io.BytesIO(template_content))
        ws = wb.active
        header_row_idx, header = _find_header_row(ws, '상품주문번호')
        header = _ensure_naver_delivery_columns(ws, header_row_idx, header)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        header_row_idx = 1
        header = _read_template_header(template_content, template_name) or list(NAVER_DELIVERY_COLUMNS)
        header = _ensure_naver_delivery_columns(ws, header_row_idx, header)
        for col_idx, col_name in enumerate(header, start=1):
            ws.cell(row=header_row_idx, column=col_idx, value=col_name)

    style_row_idx = header_row_idx + 1 if ws.max_row > header_row_idx else None
    style_by_col = {}
    if style_row_idx:
        for col_idx in range(1, len(header) + 1):
            style_by_col[col_idx] = copy(ws.cell(row=style_row_idx, column=col_idx)._style)

    if ws.max_row > header_row_idx:
        ws.delete_rows(header_row_idx + 1, ws.max_row - header_row_idx)

    column_indexes = {
        col_name: header.index(_find_naver_delivery_header(header, col_name)) + 1
        for col_name in NAVER_DELIVERY_COLUMNS
    }
    for row_offset, row_data in enumerate(rows, start=1):
        row_idx = header_row_idx + row_offset
        for col_name in NAVER_DELIVERY_COLUMNS:
            col_idx = column_indexes[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
            if col_idx in style_by_col:
                cell._style = copy(style_by_col[col_idx])
            if col_name in ('상품주문번호', '송장번호'):
                cell.number_format = '@'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def create_naver_delivery_file(file_content, file_name, invoice_map, template_content=None, template_name=None):
    df = _read_naver_order_df(file_content, file_name)
    if df is None or df.empty:
        return None

    product_order_col = '상품주문번호'
    order_col = pick_first_col(df.columns, ['주문번호', '고객주문번호'])

    rows = []
    for _, row in df.iterrows():
        product_order_no = normalize_excel_id(row.get(product_order_col))
        if not product_order_no or product_order_no.lower() == 'nan':
            continue

        order_no = normalize_excel_id(row.get(order_col)) if order_col else ""
        invoice = invoice_map.get(product_order_no) or invoice_map.get(order_no, "")
        rows.append({
            '상품주문번호': product_order_no,
            '배송방법': NAVER_DELIVERY_METHOD,
            '택배사': NAVER_DELIVERY_COMPANY,
            '송장번호': invoice
        })

    if not rows:
        return None

    return _write_naver_delivery_xlsx(
        rows,
        template_content=template_content,
        template_name=template_name
    )

def add_invoice_to_coupang(file_content, file_name, invoice_map):
    """쿠팡 파일에 운송장번호 추가 (서식 유지)"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        
        # 주문번호와 운송장번호 컬럼 찾기
        try:
            order_col_idx = header.index('주문번호') + 1
        except:
            return None
        
        # 운송장번호 컬럼이 있는지 확인
        if '운송장번호' in header:
            invoice_col_idx = header.index('운송장번호') + 1
        else:
            # 없으면 맨 끝에 추가
            invoice_col_idx = len(header) + 1
            ws.cell(row=1, column=invoice_col_idx, value='운송장번호')
        
        # 데이터 행에 운송장번호 추가
        for row_idx in range(2, ws.max_row + 1):
            order_no = normalize_excel_id(ws.cell(row=row_idx, column=order_col_idx).value)
            invoice = invoice_map.get(order_no, '')
            
            cell = ws.cell(row=row_idx, column=invoice_col_idx)
            cell.value = invoice
            # 숫자를 텍스트로 저장하여 E 표기 방지
            if invoice:
                cell.number_format = '@'  # 텍스트 형식

        _set_text_format_for_columns(
            ws,
            header,
            keyword_cols=['전화', '연락처', '휴대폰']
        )
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    except Exception as e:
        st.warning(f"쿠팡 정렬 중 오류: {e}")
        return None

def _split_paste_line(line):
    if '\t' in line:
        return [c.strip() for c in line.split('\t')]
    if ',' in line:
        return [c.strip() for c in line.split(',')]
    return [line.strip()]

def parse_pasted_sales(text, normalize=True):
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if not lines:
        return pd.DataFrame(columns=['상품명', '수량']), 0

    name_idx = None
    qty_idx = None
    start_idx = 0

    header_cols = _split_paste_line(lines[0])
    for idx, col in enumerate(header_cols):
        col_str = str(col)
        if any(k in col_str for k in ['상품', '품목']):
            name_idx = idx
        if '수량' in col_str:
            qty_idx = idx
    if name_idx is not None and qty_idx is not None:
        start_idx = 1

    parsed = []
    for line in lines[start_idx:]:
        cols = _split_paste_line(line)

        name = ""
        qty_str = ""

        if name_idx is not None and qty_idx is not None and len(cols) > max(name_idx, qty_idx):
            name = cols[name_idx]
            qty_str = cols[qty_idx]
        elif len(cols) >= 2:
            name = cols[0]
            for c in cols[1:]:
                if re.search(r'\d', c):
                    qty_str = c
                    break
        else:
            match = re.search(r'(\d+)\s*$', line)
            if match:
                qty_str = match.group(1)
                name = line[:match.start()].strip()

        if not name:
            continue

        qty_val = int(re.sub(r'[^0-9]', '', str(qty_str)) or 0)
        if qty_val <= 0:
            continue

        raw_name = str(name).strip()
        name_parts = [n.strip() for n in raw_name.split(',') if n.strip()]
        if not name_parts:
            continue

        if len(name_parts) == 1:
            final_name = identify_product(name_parts[0]) if normalize else name_parts[0]
            parsed.append({'상품명': final_name, '수량': qty_val})
        else:
            # Split total quantity across items (e.g., 4 items with qty 4 -> 1 each)
            base_qty = qty_val // len(name_parts)
            remainder = qty_val % len(name_parts)
            for idx, part in enumerate(name_parts):
                part_qty = base_qty + (1 if idx < remainder else 0)
                if part_qty <= 0:
                    continue
                final_name = identify_product(part) if normalize else part
                parsed.append({'상품명': final_name, '수량': part_qty})

    if not parsed:
        return pd.DataFrame(columns=['상품명', '수량']), 0

    df = pd.DataFrame(parsed)
    summary = df.groupby('상품명')['수량'].sum().reset_index()
    total_qty = int(summary['수량'].sum())
    return summary, total_qty

def process_data(file_name, content):
    market_key = 'unknown'
    config = {}
    for k, v in MARKET_CONFIG.items():
        if v['key'] in file_name:
            market_key = k
            config = v
            break

    if market_key == 'unknown':
        # 파일명으로 매칭되지 않는 경우 컬럼 기반 탐지 시도 (11번가 주문시트 등)
        try:
            df_probe = pd.read_csv(io.BytesIO(content)) if file_name.endswith('.csv') \
                else pd.read_excel(io.BytesIO(content))
            detected = detect_market_by_columns(df_probe)
            if detected:
                market_key = detected
                config = MARKET_CONFIG[detected]
            else:
                # 11번가 주문시트가 상단에 안내 행이 있는 경우를 위한 추가 시도
                df_probe = pd.read_csv(io.BytesIO(content), skiprows=2) if file_name.endswith('.csv') \
                    else pd.read_excel(io.BytesIO(content), skiprows=2)
                detected = detect_market_by_columns(df_probe)
                if detected:
                    market_key = detected
                    config = dict(MARKET_CONFIG[detected])
                    config['skip'] = 2
        except Exception:
            pass

    if market_key == 'unknown':
        return pd.DataFrame()

    try:
        df = pd.read_csv(io.BytesIO(content), skiprows=config.get('skip', 0)) if file_name.endswith('.csv') \
             else pd.read_excel(io.BytesIO(content), skiprows=config.get('skip', 0))

        # 11번가 주문시트는 파일명 매칭이 되더라도 헤더 위치가 다를 수 있어 재시도
        if market_key in ['11st', '11st_manual']:
            required_11st = {'주문번호', '주소', '상품명', '수량'}
            if not required_11st.issubset(set(df.columns.astype(str))):
                df_retry = pd.read_csv(io.BytesIO(content), skiprows=2) if file_name.endswith('.csv') \
                    else pd.read_excel(io.BytesIO(content), skiprows=2)
                if required_11st.issubset(set(df_retry.columns.astype(str))):
                    df = df_retry

        if market_key == 'naver':
            df['final_msg'] = df.apply(lambda r: get_message(r, ['배송메세지', '비고']), axis=1)
            mapped = pd.DataFrame({
                '고객주문번호': df['주문번호'].astype(str),
                '받는분성명': df['수취인명'],
                '받는분전화번호': df['수취인연락처1'].apply(clean_phone),
                '받는분주소': df['통합배송지'],
                '배송메세지': df['final_msg'],
                '품목': df.apply(lambda r: code_to_item(r.get('판매자 상품코드')) or identify_product(r.get('상품명', '')), axis=1),
                '수량': df['수량'],
                '내부정렬키': df['상품명'].astype(str)
            })
        elif market_key == 'coupang':
            df['final_msg'] = df.apply(lambda r: get_message(r, ['배송메세지', '비고']), axis=1)
            mapped = pd.DataFrame({
                '고객주문번호': df['주문번호'].astype(str),
                '받는분성명': df['수취인이름'],
                '받는분전화번호': df['수취인전화번호'].apply(clean_phone),
                '받는분주소': df['수취인 주소'],
                '배송메세지': df['final_msg'],
                '품목': df.apply(lambda r: code_to_item(r.get('업체상품코드')) or identify_product(r.get('등록상품명', '')), axis=1),
                '수량': df['구매수(수량)'],
                '내부정렬키': df['업체상품코드'].astype(str)
            })
        elif market_key == 'esm':
            df['final_msg'] = df.apply(lambda r: get_message(r, ['배송시 요구사항', '배송메세지', '비고']), axis=1)
            mapped = pd.DataFrame({
                '고객주문번호': df['주문번호'].astype(str),
                '받는분성명': df['수령인명'],
                '받는분전화번호': df['수령인 휴대폰'].apply(clean_phone),
                '받는분주소': df['주소'],
                '배송메세지': df['final_msg'],
                '품목': df['상품명'].apply(identify_product),
                '수량': df['수량'],
                '내부정렬키': df['상품명'].astype(str)
            })
        elif market_key in ['11st', '11st_manual']:
            name_col = '수취인' if '수취인' in df.columns else '받는분'
            phone_col = '휴대폰번호' if '휴대폰번호' in df.columns else (
                '수취인연락처' if '수취인연락처' in df.columns else '전화번호'
            )
            df['final_msg'] = df.apply(lambda r: get_message(r, ['배송메시지', '배송메세지', '비고']), axis=1)
            mapped = pd.DataFrame({
                '고객주문번호': df['주문번호'].astype(str),
                '받는분성명': df[name_col],
                '받는분전화번호': df[phone_col].apply(clean_phone),
                '받는분주소': df['주소'],
                '배송메세지': df['final_msg'],
                '품목': df['상품명'].apply(identify_product),
                '수량': df['수량'],
                '내부정렬키': df['상품명'].astype(str)
            })
        elif market_key == 'wadiz':
            df['final_msg'] = df.apply(lambda r: get_message(r, ['배송 요청 사항', '주문 요청 사항']), axis=1)
            mapped = pd.DataFrame({
                '고객주문번호': df['주문 번호'].astype(str),
                '받는분성명': df['받는 분'],
                '받는분전화번호': df['받는 분 연락처'].apply(clean_phone),
                '받는분주소': df['배송지 주소'],
                '배송메세지': df['final_msg'],
                '품목': df['주문 상품'].apply(identify_product),
                '수량': df['주문 수량'],
                '내부정렬키': df['주문 상품'].astype(str)
            })
        elif market_key == 'own':
            df['final_msg'] = df.apply(lambda r: get_message(r, ['비고', '배송메세지']), axis=1)
            mapped = pd.DataFrame({
                '고객주문번호': df['주문번호'].astype(str),
                '받는분성명': df['수령인'],
                '받는분전화번호': df['핸드폰'].apply(clean_phone),
                '받는분주소': df['주소'],
                '배송메세지': df['final_msg'],
                '품목': df['주문상품명'].apply(identify_product),
                '수량': df['수량'],
                '내부정렬키': df['주문상품명'].astype(str)
            })
        else:
            return pd.DataFrame()

        mapped['마켓순서'] = config['order']
        return mapped
    except Exception as e:
        st.error(f"❌ {file_name} 처리 실패: {e}")
        return pd.DataFrame()

def consolidate(group):
    prod_counts = group.groupby('품목')['수량'].sum().reset_index()
    def sort_key(item):
        order = {'IH_RE': 0, 'OH': 1, 'OH_RE': 2, 'PH': 3, 'PH_RE': 4, 'SH': 5, 'SH_RE': 6}
        return (order.get(str(item).upper(), 7), str(item))

    formatted = [f"{row['품목']} {int(row['수량'])}개" if row['수량'] > 1 else str(row['품목']) 
                 for _, row in prod_counts.iterrows()]
    formatted.sort(key=lambda x: sort_key(x.split(' ')[0]))

    non_empty_msgs = group['배송메세지'][group['배송메세지'] != ""].unique()
    final_msg = non_empty_msgs[0] if len(non_empty_msgs) > 0 else ""

    return {
        '고객주문번호': group.iloc[0]['고객주문번호'],
        '받는분성명': group.iloc[0]['받는분성명'],
        '받는분전화번호': group.iloc[0]['받는분전화번호'],
        '받는분주소': group.iloc[0]['받는분주소'],
        '배송메세지': final_msg,
        '품목명': ", ".join(formatted),
        '기타1': group['수량'].sum(),
        '마켓순서': group.iloc[0]['마켓순서'],
        '최종정렬키': group['내부정렬키'].min()
    }

# ==========================================
# Streamlit UI
# ==========================================
st.title("📦 자동 발주 파일 생성기")
st.markdown("---")

# 세션 상태 초기화
if 'generated_file' not in st.session_state:
    st.session_state.generated_file = None
if 'coupang_file' not in st.session_state:
    st.session_state.coupang_file = None
if 'file_info' not in st.session_state:
    st.session_state.file_info = None
if 'preview_data' not in st.session_state:
    st.session_state.preview_data = None
if 'order_mgmt_file' not in st.session_state:
    st.session_state.order_mgmt_file = None
if 'order_mgmt_info' not in st.session_state:
    st.session_state.order_mgmt_info = None
if 'order_mgmt_preview' not in st.session_state:
    st.session_state.order_mgmt_preview = None
if 'order_mgmt_raw_data' not in st.session_state:
    st.session_state.order_mgmt_raw_data = None
if 'coupang_delivery_file' not in st.session_state:
    st.session_state.coupang_delivery_file = None
if 'naver_delivery_file' not in st.session_state:
    st.session_state.naver_delivery_file = None
if 'uploaded_market_files' not in st.session_state:
    st.session_state.uploaded_market_files = None

# 사용법 안내
with st.expander("📖 사용법", expanded=False):
    st.markdown("""
    ### 📦 발주 파일 생성
    **파일 준비**
    - **네이버 파일**: 암호가 있는 경우 먼저 제거해주세요
      - 엑셀 파일 열기 → F12 → 도구 → 일반 옵션 → 비밀번호 삭제 → 저장
    
    **사용 순서**
    1. 아래 "📂 파일 업로드"에서 각 마켓의 발주 파일을 업로드하세요 (여러 개 동시 선택 가능)
    2. **파일 생성** 버튼을 클릭하세요
    3. 생성된 파일을 다운로드하세요 (여러 번 가능)
       - `MMDD_HH.xlsx`: CJ택배 업로드용 통합 발주 파일
       - `MMDD_HH_쿠팡_원본정렬.xlsx`: 쿠팡 파일 정렬본 (쿠팡 파일이 있는 경우)
    4. 새로운 파일을 처리하려면 **초기화** 버튼을 누르고 다시 시작하세요
    
    ---
    
    ### 📋 주문관리 시트 생성
    **파일 준비**
    - **CJ택배 파일**: CJ 배송 실적 출력 파일 (운송장번호와 고객주문번호 포함, 여러 파일 가능)
    - **마켓 주문 파일**: 각 마켓의 주문 내역 파일 (위에서 업로드한 파일 재사용 가능)
    
    **사용 순서**
    1. "📋 주문관리 시트 생성" 섹션으로 이동하세요
    2. CJ택배 파일을 업로드하세요
    3. 마켓 주문 파일을 업로드하세요
       - 위에서 이미 업로드했다면 "위에서 업로드한 파일 재사용" 체크박스 선택
    4. **주문관리시트 생성** 버튼을 클릭하세요
    5. 생성된 파일을 다운로드하세요
       - `주문관리_MMDD_HH.xlsx`: 송장번호 매칭된 통합 주문 관리 시트
       - `쿠팡발송_MMDD_HH.xlsx`: 쿠팡 발송용 파일 (원본 서식 유지 + 운송장번호)
       - `네이버발송_MMDD_HH.xlsx`: 네이버 엑셀발송용 파일 (상품주문번호 + CJ 송장번호)
    6. 데이터 미리보기와 품목별 판매 집계를 확인하세요
    
    **자동 기능**
    - ✅ 같은 주문번호의 제품 자동 통합 (예: OH 2개, PH 1개 → 한 줄로 표시)
    - ✅ CJ 고객주문번호와 매칭되는 송장번호 자동 입력
    - ✅ 발주파일과 동일한 순서로 자동 정렬 (마켓별 → 제품별)
    - ✅ 옥션/지마켓 자동 구분 (주문번호 패턴 분석)
    - ✅ 제품명 자동 분류 (IH_Re, OH, OH_Re, PH, PH_Re, SH, SH_Re, 케이블, 거치대, 번호판 등)
    - ✅ 쿠팡 발송 파일 자동 생성 (원본 서식 유지, 운송장번호 추가)
    - ✅ 네이버 엑셀발송 파일 자동 생성 (배송방법: 택배,등기,소포 / 택배사: CJ대한통운)

     ---

     ### 📊 품목별 판매 집계
     1. 상품명/수량을 복사해서 붙여넣기
         - 탭/콤마 구분 자동 인식

     2. 집계 옵션 선택
         - 상품명 자동 분류 적용: OH/PH/SH, 케이블(일반), 케이블s 등으로 분류
         - 붙여넣기 즉시 자동 집계: 해제 시 "집계하기" 버튼으로 실행
    
    ---
    
    ### 📌 지원 마켓
    - 네이버 스마트스토어
    - 쿠팡 (DeliveryList)
    - 와디즈 (발송 처리용 주문)
    - 자사몰 (orders)
    - ESM (지마켓/옥션 - 신규주문)
    - 11번가 (allList)

    ### 📦 지원 상품
    - 기존 제품과 `_Re` 리퍼 제품을 각각 별도 품목으로 처리
    
    ### 💡 참고사항
    - 파일명 시간 형식: MMDD_HH (예: 0206_15 = 2월 6일 오후 3시)
    - 동일한 배송지로 여러 상품 주문 시 자동 통합
    - 정렬 순서: 네이버→쿠팡→자사몰→ESM→11번가→와디즈 / IH_Re→OH→OH_Re→PH→PH_Re→SH→SH_Re→기타
    """)

st.markdown("### 📂 파일 업로드")

# 초기화 버튼 (생성된 파일이 있을 때만 표시)
if st.session_state.generated_file:
    if st.button("🔄 초기화 (새로운 파일 처리)", type="secondary"):
        st.session_state.generated_file = None
        st.session_state.coupang_file = None
        st.session_state.file_info = None
        st.session_state.preview_data = None
        st.rerun()

uploaded_files = st.file_uploader(
    "발주 파일을 선택하세요 (여러 파일 선택 가능)",
    type=['csv', 'xlsx', 'xls'],
    accept_multiple_files=True,
    help="네이버, 쿠팡, 자사몰, ESM, 11번가 등의 발주 파일을 모두 선택하세요",
    disabled=st.session_state.generated_file is not None
)

if uploaded_files and not st.session_state.generated_file:
    st.success(f"✅ {len(uploaded_files)}개 파일 업로드됨")
    
    # 세션에 파일 저장 (주문관리시트에서 재사용 가능)
    st.session_state.uploaded_market_files = [(f.name, f.read()) for f in uploaded_files]
    
    # 업로드된 파일 목록 표시
    with st.expander("업로드된 파일 목록"):
        for file in uploaded_files:
            st.write(f"- {file.name}")

if st.button("🚀 발주 파일 생성", type="primary", disabled=not uploaded_files or st.session_state.generated_file is not None):
    with st.spinner("파일 처리 중..."):
        combined_list = []
        coupang_sorted = None
        
        now = datetime.now(ZoneInfo("Asia/Seoul"))
        date_prefix = now.strftime('%m%d')
        time_suffix = now.strftime('%H')
        
        # 파일 처리 (세션에 저장된 파일 사용)
        for file_name, content in st.session_state.uploaded_market_files:
            
            # 쿠팡 파일인 경우 정렬된 버전 생성
            if 'DeliveryList' in file_name:
                coupang_sorted = sort_xlsx_preserving_format(content, '업체상품코드')
                if coupang_sorted:
                    coupang_sorted = apply_text_format_to_excel_bytes(
                        coupang_sorted,
                        keyword_cols=['전화', '연락처', '휴대폰']
                    )
            
            # 데이터 처리
            temp_df = process_data(file_name, content)
            if not temp_df.empty:
                combined_list.append(temp_df)
        
        if combined_list:
            # 데이터 병합 및 처리
            full_df = pd.concat(combined_list, ignore_index=True)
            
            final_data = []
            groups = full_df.groupby(['받는분성명', '받는분전화번호', '받는분주소'], sort=False)
            for name, group in groups:
                final_data.append(consolidate(group))
            
            final_df = pd.DataFrame(final_data)
            final_df = final_df.sort_values(by=['마켓순서', '최종정렬키'])
            
            # 최종 파일 생성
            final_filename = f"{date_prefix}_{time_suffix}.xlsx"
            final_cols = ['고객주문번호', '받는분성명', '받는분전화번호', '받는분주소(전체, 분할)', '배송메세지1', '품목명', '기타1']
            
            output = io.BytesIO()
            final_df.rename(columns={
                '받는분주소': '받는분주소(전체, 분할)',
                '배송메세지': '배송메세지1'
            }).to_excel(output, index=False, columns=final_cols)
            output.seek(0)
            formatted_order_file = apply_text_format_to_excel_bytes(
                output.getvalue(),
                target_cols=['받는분전화번호'],
                keyword_cols=['전화', '연락처', '휴대폰']
            )
            
            # 세션 상태에 저장
            st.session_state.generated_file = formatted_order_file
            st.session_state.coupang_file = coupang_sorted
            st.session_state.file_info = {
                'filename': final_filename,
                'coupang_filename': f"{date_prefix}_{time_suffix}_쿠팡_원본정렬.xlsx",
                'order_count': len(final_df)
            }
            st.session_state.preview_data = final_df[['고객주문번호', '받는분성명', '품목명', '기타1']]
            
            st.success("✅ 발주 파일 생성 완료!")
            st.rerun()
        else:
            st.error("❌ 처리할 수 있는 파일이 없습니다. 파일 형식을 확인해주세요.")

# 생성된 파일이 있으면 다운로드 섹션 표시
if st.session_state.generated_file:
    st.markdown("---")
    st.markdown("### 📥 파일 다운로드")
    st.info("💡 아래 버튼을 원하는 만큼 클릭하여 파일을 다운로드하세요. 다운로드 후에도 파일은 유지됩니다.")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.download_button(
            label="📄 발주 파일 다운로드",
            data=st.session_state.generated_file,
            file_name=st.session_state.file_info['filename'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    if st.session_state.coupang_file:
        with col2:
            st.download_button(
                label="📄 쿠팡 정렬 파일 다운로드",
                data=st.session_state.coupang_file,
                file_name=st.session_state.file_info['coupang_filename'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    # 미리보기
    with st.expander("📊 데이터 미리보기", expanded=True):
        st.dataframe(st.session_state.preview_data, use_container_width=True)
        st.info(f"총 주문 건수: {st.session_state.file_info['order_count']}건")


# 주문관리시트 생성 섹션 추가 코드

# Footer 앞에 추가할 내용:

st.markdown("---")
st.markdown("## 📋 주문관리시트 생성 (송장번호 매칭)")
st.markdown("발주 후 CJ택배에서 받은 송장번호 파일과 마켓 주문시트를 매칭하여 주문관리시트를 생성합니다.")

# 세션 상태 초기화 (주문관리시트용)
if 'order_mgmt_file' not in st.session_state:
    st.session_state.order_mgmt_file = None
if 'order_mgmt_info' not in st.session_state:
    st.session_state.order_mgmt_info = None
if 'naver_delivery_file' not in st.session_state:
    st.session_state.naver_delivery_file = None

col_a, col_b = st.columns(2)

with col_a:
    cj_files = st.file_uploader(
        "CJ택배 출력 파일 업로드",
        type=['xlsx', 'xls', 'csv'],
        key="cj_upload",
        help="운송장번호와 고객주문번호가 포함된 CJ택배 출력 파일",
        accept_multiple_files=True
    )

with col_b:
    market_files = st.file_uploader(
        "마켓 주문시트 업로드",
        type=['xlsx', 'xls', 'csv'],
        accept_multiple_files=True,
        key="market_upload",
        help="네이버, 쿠팡, 11번가 등 마켓 주문시트"
    )
    
    use_existing = st.checkbox(
        "위에서 업로드한 파일 사용하기",
        value=False,
        disabled=not st.session_state.uploaded_market_files,
        help="발주 파일 생성에서 업로드한 마켓 주문시트를 재사용합니다"
    )
    
    if use_existing and st.session_state.uploaded_market_files:
        st.info(f"✅ {len(st.session_state.uploaded_market_files)}개의 업로드된 파일 사용")
        with st.expander("사용할 파일 목록"):
            for file_name, _ in st.session_state.uploaded_market_files:
                st.write(f"- {file_name}")
        market_files = None

naver_template_file = st.file_uploader(
    "네이버 엑셀발송 양식 업로드 (선택)",
    type=['xlsx', 'xls'],
    key="naver_template_upload",
    help=f"업로드하지 않으면 sample_data/{NAVER_DELIVERY_TEMPLATE_NAME} 규격을 자동으로 사용합니다."
)

if st.button("🔗 주문관리시트 생성", type="primary", key="gen_order_mgmt"):
    if not cj_files:
        st.error("CJ택배 파일을 업로드해주세요")
    elif not use_existing and not market_files:
        st.error("마켓 주문시트를 업로드하거나 위의 파일을 사용하도록 체크해주세요")
    else:
        with st.spinner("주문관리시트 생성 중..."):
            try:
                # CJ택배 파일 읽기
                cj_dfs = []
                for cj_file in cj_files:
                    cj_content = cj_file.read()
                    cj_df = pd.read_csv(io.BytesIO(cj_content)) if cj_file.name.endswith('.csv') \
                        else pd.read_excel(io.BytesIO(cj_content))
                    cj_df.columns = cj_df.columns.astype(str).str.strip()
                    cj_dfs.append(cj_df)

                if not cj_dfs:
                    st.error("CJ택배 파일을 업로드해주세요")
                    st.stop()

                cj_df = pd.concat(cj_dfs, ignore_index=True)
                
                # 운송장번호와 고객주문번호 매핑
                invoice_map = {}
                if '운송장번호' in cj_df.columns and '고객주문번호' in cj_df.columns:
                    for _, row in cj_df.iterrows():
                        order_no = normalize_excel_id(row['고객주문번호'])
                        invoice = normalize_excel_id(row['운송장번호'])
                        if order_no and invoice and invoice != 'nan':
                            invoice_map[order_no] = invoice
                
                today_str = datetime.now(ZoneInfo("Asia/Seoul")).strftime('%Y.%m.%d')

                # 마켓 주문시트 처리
                all_orders = []
                
                # 사용할 파일 결정
                files_to_process = []
                if use_existing and st.session_state.uploaded_market_files:
                    files_to_process = st.session_state.uploaded_market_files
                else:
                    files_to_process = [(f.name, f.read()) for f in market_files]
                
                for file_name, content in files_to_process:
                    
                    # 마켓별 상세 데이터 추출
                    market_key = 'unknown'
                    config = {}
                    for k, v in MARKET_CONFIG.items():
                        if v['key'] in file_name:
                            market_key = k
                            config = v
                            break
                    
                    # 컬럼 기반 탐지
                    if market_key == 'unknown':
                        try:
                            df_probe = pd.read_csv(io.BytesIO(content)) if file_name.endswith('.csv') \
                                else pd.read_excel(io.BytesIO(content))
                            detected = detect_market_by_columns(df_probe)
                            if detected:
                                market_key = detected
                                config = MARKET_CONFIG[detected]
                            else:
                                df_probe = pd.read_csv(io.BytesIO(content), skiprows=2) if file_name.endswith('.csv') \
                                    else pd.read_excel(io.BytesIO(content), skiprows=2)
                                detected = detect_market_by_columns(df_probe)
                                if detected:
                                    market_key = detected
                                    config = dict(MARKET_CONFIG[detected])
                                    config['skip'] = 2
                        except Exception:
                            pass
                    
                    if market_key == 'unknown':
                        continue
                    
                    # 데이터 읽기
                    df = pd.read_csv(io.BytesIO(content), skiprows=config.get('skip', 0)) if file_name.endswith('.csv') \
                        else pd.read_excel(io.BytesIO(content), skiprows=config.get('skip', 0))
                    df.columns = df.columns.astype(str).str.strip()
                    
                    # 11번가 헤더 재시도
                    if market_key in ['11st', '11st_manual']:
                        required_11st = {'주문번호', '주소', '상품명', '수량'}
                        if not required_11st.issubset(set(df.columns.astype(str))):
                            df_retry = pd.read_csv(io.BytesIO(content), skiprows=2) if file_name.endswith('.csv') \
                                else pd.read_excel(io.BytesIO(content), skiprows=2)
                            if required_11st.issubset(set(df_retry.columns.astype(str))):
                                df = df_retry
                                df.columns = df.columns.astype(str).str.strip()
                    
                    # 마켓별 데이터 추출
                    channel_name = {
                        'naver': '네이버',
                        'coupang': '쿠팡',
                        'own': '자사몰',
                        'esm': '지마켓',
                        '11st': '11번가',
                        '11st_manual': '11번가',
                        'wadiz': '와디즈'
                    }.get(market_key, '기타')
                    
                    if market_key == 'naver':
                        date_col = pick_first_col(df.columns, ['결제일', '주문일', '결제일시', '주문일시'])
                        buyer_col = pick_first_col(df.columns, ['구매자명', '주문자명', '구매자', '주문자'])
                        df['final_msg'] = df.apply(lambda r: get_message(r, ['배송메세지', '비고']), axis=1)
                        
                        for _, row in df.iterrows():
                            order_no = normalize_excel_id(row['주문번호'])
                            all_orders.append({
                                '날짜': today_str,
                                '채널': channel_name,
                                '주문번호': order_no,
                                '상품명': code_to_item(row.get('판매자 상품코드')) or identify_product(row.get('상품명', '')),
                                '상품명_원문': str(row.get('상품명', '')).strip(),
                                '수량': row.get('수량', ''),
                                '주문인': row.get(buyer_col, '') if buyer_col else '',
                                '수취인': row.get('수취인명', ''),
                                '전화번호': clean_phone(row.get('수취인연락처1', '')),
                                '주소': row.get('통합배송지', ''),
                                '비고': row.get('final_msg', ''),
                                '송장번호': invoice_map.get(order_no, '')
                            })
                    
                    elif market_key == 'coupang':
                        date_col = pick_first_col(df.columns, ['주문일', '결제완료시각', '결제일시', '주문일시'])
                        buyer_col = pick_first_col(df.columns, ['주문자명', '구매자', '주문자', '구매자명'])
                        df['final_msg'] = df.apply(lambda r: get_message(r, ['배송메세지', '비고']), axis=1)
                        
                        for _, row in df.iterrows():
                            order_no = normalize_excel_id(row['주문번호'])
                            all_orders.append({
                                '날짜': today_str,
                                '채널': channel_name,
                                '주문번호': order_no,
                                '상품명': code_to_item(row.get('업체상품코드')) or identify_product(row.get('등록상품명', '')),
                                '상품명_원문': str(row.get('등록상품명', '')).strip(),
                                '수량': row.get('구매수(수량)', ''),
                                '주문인': row.get(buyer_col, '') if buyer_col else '',
                                '수취인': row.get('수취인이름', ''),
                                '전화번호': clean_phone(row.get('수취인전화번호', '')),
                                '주소': row.get('수취인 주소', ''),
                                '비고': row.get('final_msg', ''),
                                '송장번호': invoice_map.get(order_no, '')
                            })
                    
                    elif market_key == 'esm':
                        date_col = pick_first_col(df.columns, ['결제일시', '주문일', '결제일', '주문일시'])
                        buyer_col = pick_first_col(df.columns, ['주문자명', '구매자명', '주문자', '구매자'])
                        df['final_msg'] = df.apply(lambda r: get_message(r, ['배송시 요구사항', '배송메세지', '비고']), axis=1)
                        
                        for _, row in df.iterrows():
                            order_no = normalize_excel_id(row['주문번호'])
                            
                            # 주문번호 패턴으로 옥션/지마켓 구분
                            if len(order_no) == 10:
                                if order_no.startswith('2'):
                                    actual_channel = '옥션'
                                elif order_no.startswith('4'):
                                    actual_channel = '지마켓'
                                else:
                                    actual_channel = channel_name
                            else:
                                actual_channel = channel_name
                            
                            all_orders.append({
                                '날짜': today_str,
                                '채널': actual_channel,
                                '주문번호': order_no,
                                '상품명': identify_product(row.get('상품명', '')),
                                '상품명_원문': str(row.get('상품명', '')).strip(),
                                '수량': row.get('수량', ''),
                                '주문인': row.get(buyer_col, '') if buyer_col else '',
                                '수취인': row.get('수령인명', ''),
                                '전화번호': clean_phone(row.get('수령인 휴대폰', '')),
                                '주소': row.get('주소', ''),
                                '비고': row.get('final_msg', ''),
                                '송장번호': invoice_map.get(order_no, '')
                            })
                    
                    elif market_key in ['11st', '11st_manual']:
                        date_col = pick_first_col(df.columns, ['결제일시', '주문일', '결제일', '주문일시'])
                        buyer_col = pick_first_col(df.columns, ['구매자', '주문자', '구매자명', '주문자명'])
                        name_col = pick_first_col(df.columns, ['수취인', '받는분'])
                        phone_col = pick_first_col(df.columns, ['휴대폰번호', '수취인연락처', '전화번호'])
                        df['final_msg'] = df.apply(lambda r: get_message(r, ['배송메시지', '배송메세지', '비고']), axis=1)
                        
                        for _, row in df.iterrows():
                            order_no = normalize_excel_id(row['주문번호'])
                            all_orders.append({
                                '날짜': today_str,
                                '채널': channel_name,
                                '주문번호': order_no,
                                '상품명': identify_product(row.get('상품명', '')),
                                '상품명_원문': str(row.get('상품명', '')).strip(),
                                '수량': row.get('수량', ''),
                                '주문인': row.get(buyer_col, '') if buyer_col else '',
                                '수취인': row.get(name_col, '') if name_col else '',
                                '전화번호': clean_phone(row.get(phone_col, '')) if phone_col else '',
                                '주소': row.get('주소', ''),
                                '비고': row.get('final_msg', ''),
                                '송장번호': invoice_map.get(order_no, '')
                            })
                    
                    elif market_key == 'own':
                        date_col = pick_first_col(df.columns, ['주문일시', '주문일', '결제일', '결제일시'])
                        buyer_col = pick_first_col(df.columns, ['주문자', '구매자', '주문자명', '구매자명'])
                        df['final_msg'] = df.apply(lambda r: get_message(r, ['비고', '배송메세지']), axis=1)
                        
                        for _, row in df.iterrows():
                            order_no = normalize_excel_id(row['주문번호'])
                            all_orders.append({
                                '날짜': today_str,
                                '채널': channel_name,
                                '주문번호': order_no,
                                '상품명': identify_product(row.get('주문상품명', '')),
                                '상품명_원문': str(row.get('주문상품명', '')).strip(),
                                '수량': row.get('수량', ''),
                                '주문인': row.get(buyer_col, '') if buyer_col else '',
                                '수취인': row.get('수령인', ''),
                                '전화번호': clean_phone(row.get('핸드폰', '')),
                                '주소': row.get('주소', ''),
                                '비고': row.get('final_msg', ''),
                                '송장번호': invoice_map.get(order_no, '')
                            })
                    elif market_key == 'wadiz':
                        buyer_col = pick_first_col(df.columns, ['서포터 이름', '주문자', '구매자', '주문자명', '구매자명'])
                        df['final_msg'] = df.apply(lambda r: get_message(r, ['배송 요청 사항', '주문 요청 사항']), axis=1)

                        for _, row in df.iterrows():
                            order_no = normalize_excel_id(row.get('주문 번호', ''))
                            all_orders.append({
                                '날짜': today_str,
                                '채널': channel_name,
                                '주문번호': order_no,
                                '상품명': identify_product(row.get('주문 상품', '')),
                                '상품명_원문': str(row.get('주문 상품', '')).strip(),
                                '수량': row.get('주문 수량', ''),
                                '주문인': row.get(buyer_col, '') if buyer_col else '',
                                '수취인': row.get('받는 분', ''),
                                '전화번호': clean_phone(row.get('받는 분 연락처', '')),
                                '주소': row.get('배송지 주소', ''),
                                '비고': row.get('final_msg', ''),
                                '송장번호': invoice_map.get(order_no, '')
                            })
                
                # 주문관리시트 생성
                if all_orders:
                    mgmt_df = pd.DataFrame(all_orders)
                    
                    # 같은 주문번호로 제품 통합
                    consolidated_list = []
                    
                    for (channel, order_no), group in mgmt_df.groupby(['채널', '주문번호']):
                        # 제품별 수량 집계
                        prod_counts = {}
                        for _, row in group.iterrows():
                            prod = row['상품명']
                            qty = row['수량']
                            if prod in prod_counts:
                                prod_counts[prod] += qty
                            else:
                                prod_counts[prod] = qty
                        
                        # IH_Re, OH, OH_Re, PH, PH_Re, SH, SH_Re 순서로 정렬
                        def get_sort_priority(prod_name):
                            prod_upper = str(prod_name).strip().upper()
                            order = {'IH_RE': 0, 'OH': 1, 'OH_RE': 2, 'PH': 3, 'PH_RE': 4, 'SH': 5, 'SH_RE': 6}
                            return (order.get(prod_upper, 7), prod_name)
                        
                        sorted_prods = sorted(prod_counts.items(), key=lambda x: get_sort_priority(x[0]))
                        
                        # "OH 2개, PH 1개" 형태로 포맷팅
                        formatted = []
                        for prod, qty in sorted_prods:
                            if qty > 1:
                                formatted.append(f"{prod} {int(qty)}개")
                            else:
                                formatted.append(str(prod))
                        
                        # 첫 번째 제품으로 정렬키 결정
                        first_prod = sorted_prods[0][0] if sorted_prods else ''
                        first_prod_priority = get_sort_priority(first_prod)[0]
                        
                        # 마켓 순서 매핑
                        market_order_map = {
                            '네이버': 1,
                            '쿠팡': 2,
                            '자사몰': 3,
                            '옥션': 4,
                            '지마켓': 4,
                            '11번가': 5,
                            '와디즈': 6
                        }
                        market_order = market_order_map.get(channel, 99)
                        
                        consolidated_list.append({
                            '날짜': group.iloc[0]['날짜'],
                            '채널': channel,
                            '주문번호': order_no,
                            '상품명': ", ".join(formatted),
                            '수량': int(group['수량'].sum()),
                            '주문인': group.iloc[0]['주문인'],
                            '수취인': group.iloc[0]['수취인'],
                            '전화번호': group.iloc[0]['전화번호'],
                            '주소': group.iloc[0]['주소'],
                            '비고': group.iloc[0]['비고'],
                            '송장번호': group.iloc[0]['송장번호'],
                            '마켓순서': market_order,
                            '상품순서': first_prod_priority
                        })
                    
                    consolidated = pd.DataFrame(consolidated_list)
                    # 발주파일과 같은 순서로 정렬: 마켓 → 상품
                    consolidated = consolidated.sort_values(by=['마켓순서', '상품순서'])
                    # 정렬용 컬럼 제거
                    consolidated = consolidated.drop(columns=['마켓순서', '상품순서'])
                    
                    # 쿠팡 발송 파일 생성
                    coupang_delivery = None
                    for file_name, content in files_to_process:
                        if 'DeliveryList' in file_name:
                            coupang_delivery = add_invoice_to_coupang(content, file_name, invoice_map)
                            if coupang_delivery:
                                coupang_delivery = apply_text_format_to_excel_bytes(
                                    coupang_delivery,
                                    keyword_cols=['전화', '연락처', '휴대폰']
                                )
                            break

                    # 네이버 엑셀발송 파일 생성
                    naver_delivery = None
                    naver_template_content = None
                    naver_template_name = None

                    if naver_template_file:
                        naver_template_content = naver_template_file.read()
                        naver_template_name = naver_template_file.name
                    else:
                        local_template = find_naver_delivery_template()
                        if local_template:
                            naver_template_content = local_template.read_bytes()
                            naver_template_name = local_template.name

                    for file_name, content in files_to_process:
                        naver_delivery = create_naver_delivery_file(
                            content,
                            file_name,
                            invoice_map,
                            template_content=naver_template_content,
                            template_name=naver_template_name
                        )
                        if naver_delivery:
                            break
                    
                    # 엑셀 파일 생성
                    output = io.BytesIO()
                    consolidated.to_excel(output, index=False)
                    output.seek(0)
                    formatted_order_mgmt = apply_text_format_to_excel_bytes(
                        output.getvalue(),
                        target_cols=['전화번호'],
                        keyword_cols=['전화', '연락처', '휴대폰']
                    )
                    
                    now = datetime.now(ZoneInfo("Asia/Seoul"))
                    filename = f"주문관리_{now.strftime('%m%d_%H')}.xlsx"
                    
                    st.session_state.order_mgmt_file = formatted_order_mgmt
                    st.session_state.order_mgmt_info = {
                        'filename': filename,
                        'count': len(consolidated),
                        'matched': len(consolidated[consolidated['송장번호'] != ''])
                    }
                    st.session_state.order_mgmt_preview = consolidated
                    st.session_state.order_mgmt_raw_data = all_orders
                    st.session_state.coupang_delivery_file = coupang_delivery
                    st.session_state.naver_delivery_file = naver_delivery
                    
                    st.success("✅ 주문관리시트 생성 완료!")
                    st.rerun()
                else:
                    st.error("❌ 처리할 수 있는 주문 데이터가 없습니다.")
                    
            except Exception as e:
                st.error(f"❌ 오류 발생: {e}")

# 주문관리시트 다운로드
if st.session_state.order_mgmt_file:
    st.markdown("### 📥 주문관리시트 다운로드")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.download_button(
            label="📋 주문관리시트 다운로드",
            data=st.session_state.order_mgmt_file,
            file_name=st.session_state.order_mgmt_info['filename'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    if st.session_state.coupang_delivery_file:
        with col2:
            now = datetime.now(ZoneInfo("Asia/Seoul"))
            coupang_filename = f"쿠팡발송_{now.strftime('%m%d_%H')}.xlsx"
            st.download_button(
                label="📦 쿠팡 발송 파일 다운로드",
                data=st.session_state.coupang_delivery_file,
                file_name=coupang_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    if st.session_state.naver_delivery_file:
        with col3:
            now = datetime.now(ZoneInfo("Asia/Seoul"))
            naver_filename = f"네이버발송_{now.strftime('%m%d_%H')}.xlsx"
            st.download_button(
                label="📦 네이버 발송 파일 다운로드",
                data=st.session_state.naver_delivery_file,
                file_name=naver_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    st.info(f"총 {st.session_state.order_mgmt_info['count']}건 | 송장번호 매칭 {st.session_state.order_mgmt_info['matched']}건")
    
    # 미리보기
    with st.expander("📊 데이터 미리보기", expanded=True):
        st.dataframe(st.session_state.order_mgmt_preview, use_container_width=True)
    
    # 품목별 판매 집계
    if st.session_state.order_mgmt_raw_data:
        with st.expander("📈 품목별 판매 집계", expanded=False):
            raw_df = pd.DataFrame(st.session_state.order_mgmt_raw_data)
            use_normalized = st.checkbox(
                "상품명 자동 분류 적용 (OH/PH/SH 등)",
                value=False,
                key="mgmt_summary_normalize",
                help="체크하면 상품명을 OH/PH/SH, 케이블 등으로 자동 분류해 집계합니다"
            )

            summary_col = '상품명' if use_normalized else '상품명_원문'
            if summary_col not in raw_df.columns:
                summary_col = '상품명'

            product_summary = raw_df.groupby(summary_col)['수량'].sum().reset_index()
            product_summary.columns = ['품목', '판매 수량']
            
            # 품목 순서 정의
            product_order = {
                'IH_Re': 0,
                'OH': 1,
                'OH_Re': 2,
                'PH': 3,
                'PH_Re': 4,
                'SH': 5,
                'SH_Re': 6,
                '케이블(일반)': 7,
                '케이블s': 8,
                '휴대폰거치대': 9,
                '차량번호판': 10,
                '차량용망치': 11,
                '도막측정기': 12
            }
            
            # 정렬키 추가
            product_summary['순서'] = product_summary['품목'].map(lambda x: product_order.get(x, 99))
            product_summary = product_summary.sort_values(by=['순서', '품목'])
            product_summary = product_summary[['품목', '판매 수량']]
            
            st.dataframe(product_summary, use_container_width=True, hide_index=True)
            st.info(f"총 품목 수: {len(product_summary)}개")
    
    if st.button("🔄 새 주문관리시트 생성", key="reset_mgmt"):
        st.session_state.order_mgmt_file = None
        st.session_state.order_mgmt_info = None
        st.session_state.order_mgmt_preview = None
        st.session_state.order_mgmt_raw_data = None
        st.session_state.coupang_delivery_file = None
        st.session_state.naver_delivery_file = None
        st.rerun()

st.markdown("---")
st.markdown("## 📊 품목별 판매 집계 (복붙 입력)")
st.markdown("오전/오후 발주 후 시트에서 상품명과 수량을 복사해 붙여넣으면 품목별/하루 총 판매량을 집계합니다.")

toggle_col1, toggle_col2, _ = st.columns([1, 1, 2])
with toggle_col1:
    normalize_names = st.checkbox(
        "상품명 자동 분류 적용 (OH/PH/SH 등)",
        value=True,
        help="상품명을 OH/PH/SH, 케이블, 거치대 등으로 자동 분류합니다"
    )

with toggle_col2:
    auto_calc = st.checkbox(
        "붙여넣기 즉시 자동 집계",
        value=True,
        help="체크 해제 시 '집계하기' 버튼을 눌러야 집계됩니다"
    )

if 'paste_summary_ready' not in st.session_state:
    st.session_state.paste_summary_ready = False

def _mark_paste_ready():
    st.session_state.paste_summary_ready = True

pasted_text = st.text_area(
    "상품명과 수량을 붙여넣기",
    placeholder="예)\n상품명\t수량\nOH\t2\nPH\t1\n케이블\t3",
    height=160,
    on_change=_mark_paste_ready if auto_calc else None,
    key="paste_input"
)

col_calc, _ = st.columns([1, 3])
with col_calc:
    if st.button("집계하기", type="primary"):
        st.session_state.paste_summary_ready = True

if st.session_state.paste_summary_ready and pasted_text.strip():
    summary_df, total_qty = parse_pasted_sales(pasted_text, normalize=normalize_names)
    if summary_df.empty:
        st.warning("집계할 데이터가 없습니다. 붙여넣은 내용을 확인해주세요.")
    else:
        product_order = {
            'IH_Re': 0,
            'OH': 1,
            'OH_Re': 2,
            'PH': 3,
            'PH_Re': 4,
            'SH': 5,
            'SH_Re': 6,
            '케이블(일반)': 7,
            '케이블s': 8,
            '휴대폰거치대': 9,
            '차량번호판': 10,
            '차량용망치': 11,
            '도막측정기': 12
        }
        summary_df['순서'] = summary_df['상품명'].map(lambda x: product_order.get(x, 99))
        summary_df = summary_df.sort_values(by=['순서', '상품명']).drop(columns=['순서'])
        summary_df.columns = ['품목', '판매 수량']
        st.dataframe(summary_df, use_container_width=True, hide_index=True)
        st.info(f"하루 총 판매 수량: {total_qty}개")

# Footer
st.markdown("---")
st.markdown(
    """
    <div style='text-align: center; color: gray;'>
    자동 발주 파일 생성기 | Made by 🦖 DandiHaza
    </div>
    """,
    unsafe_allow_html=True
)
